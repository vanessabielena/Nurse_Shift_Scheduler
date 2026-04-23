import random
import json
import datetime
import calendar
from openpyxl import Workbook
import tkinter as tk
from tkinter import ttk
import sys
import os



# =====================
# KONŠTANTY
# =====================

DENNA_POCET = 6
NOCNA_POCET = 6
MESIAC = int(input("Zadaj mesiac (príklad: noveber - 11): "))
ROK = int(input("Zadaj rok: "))
MAX_SLUZIEB = int(input("Zadaj počet služieb na osobu: "))
NAHLASENE_VOLNO = {}
MAX_POKUSOV = 1000

# def resource_path(relative_path):
#     try:
#         base_path = sys._MEIPASS
#     except Exception:
#         base_path = os.path.abspath(".")
#
#     return os.path.join(base_path, relative_path)

MENA_SUBOR = "mena.txt"
JSON_SUBOR = "sluzby_mesacne.json"
EXCEL_SUBOR = "sluzby_mesacne_"+ str(MESIAC) + "_" + str(ROK) + ".xlsx"

# =====================
# FUNKCIE
# =====================

def nacitaj_mena(subor):
    with open(subor, "r", encoding="utf-8") as f:
        mena = [riadok.strip() for riadok in f if riadok.strip()]
    return mena

ZAMESTNANCI = nacitaj_mena(MENA_SUBOR)
def zisti_mesiac():
    # dnes = datetime.date.today()
    rok = ROK #dnes.year
    mesiac = MESIAC #(dnes.month + 1) % 12
    pocet_dni = calendar.monthrange(rok, mesiac)[1]
    return rok, mesiac, pocet_dni

def je_smenova_sestra(meno):
    return "s.s." in meno

def moze_dostat(meno, typ, historia):
    posledne = historia[meno][-2:]
    return not (posledne == [typ, typ])

def vytvor_rozpis(rok, mesiac, pocet_dni):
    rozpis = {}

    zostavajuce = {m: MAX_SLUZIEB for m in ZAMESTNANCI}
    historia = {m: [] for m in ZAMESTNANCI}

    nocna_vcera = set()
    smenove_sestry = [m for m in ZAMESTNANCI if je_smenova_sestra(m)]

    for den in range(1, pocet_dni + 1):
        datum = f"{rok}-{mesiac:02d}-{den:02d}"

        # =====================
        # NOČNÁ SLUŽBA
        # =====================
        kandidati_noc = [
            m for m in ZAMESTNANCI
            if zostavajuce[m] > 0
            and moze_dostat(m, "N", historia)
            and den not in NAHLASENE_VOLNO.get(m, set())
        ]

        ss_noc = [m for m in kandidati_noc if je_smenova_sestra(m)]
        if not ss_noc:
            raise ValueError(f"Chýba s.s. na nočnú službu ({datum})")

        prva_nocna = random.choice(ss_noc)
        nocna = [prva_nocna]

        zvysok_noc = [m for m in kandidati_noc if m != prva_nocna]
        if len(zvysok_noc) < NOCNA_POCET - 1:
            raise ValueError(f"Nedostatok ľudí na nočnú službu ({datum})")

        nocna += random.sample(zvysok_noc, NOCNA_POCET - 1)

        for m in nocna:
            zostavajuce[m] -= 1
            historia[m].append("N")

        # =====================
        # DENNÁ SLUŽBA
        # =====================
        kandidati_den = [
            m for m in ZAMESTNANCI
            if zostavajuce[m] > 0
            and m not in nocna
            and m not in nocna_vcera
            and moze_dostat(m, "D", historia)
            and den not in NAHLASENE_VOLNO.get(m, set())
        ]

        ss_den = [m for m in kandidati_den if je_smenova_sestra(m)]
        if not ss_den:
            raise ValueError(f"Chýba s.s. na dennú službu ({datum})")

        prva_denna = random.choice(ss_den)
        denna = [prva_denna]

        zvysok_den = [m for m in kandidati_den if m != prva_denna]
        if len(zvysok_den) < DENNA_POCET - 1:
            raise ValueError(f"Nedostatok ľudí na dennú službu ({datum})")

        denna += random.sample(zvysok_den, DENNA_POCET - 1)

        for m in denna:
            zostavajuce[m] -= 1
            historia[m].append("D")

        # =====================
        # VOĽNO
        # =====================
        for m in ZAMESTNANCI:
            if m not in denna and m not in nocna:
                historia[m].append("-")

        rozpis[datum] = {
            "denna": denna,
            "nocna": nocna
        }

        nocna_vcera = set(nocna)

    return rozpis


def spusti_ui(pocet_dni):
    global NAHLASENE_VOLNO

    root = tk.Tk()
    root.title("Nahlásenie voľna")

    vybrane_meno = tk.StringVar()
    vybrane_meno.set(ZAMESTNANCI[0])

    ttk.Label(root, text="Vyber osobu:").pack()
    ttk.Combobox(root, values=ZAMESTNANCI, textvariable=vybrane_meno).pack()

    dni_vars = {}

    frame = ttk.Frame(root)
    frame.pack()

    for d in range(1, pocet_dni + 1):
        var = tk.BooleanVar()
        chk = ttk.Checkbutton(frame, text=str(d), variable=var)
        chk.grid(row=(d-1)//7, column=(d-1)%7)
        dni_vars[d] = var

    def uloz_volno():
        meno = vybrane_meno.get()
        dni = {d for d, v in dni_vars.items() if v.get()}
        NAHLASENE_VOLNO[meno] = dni
        print(f"Uložené voľno pre {meno}: {dni}")

    def start():
        root.destroy()

    ttk.Button(root, text="Uložiť voľno", command=uloz_volno).pack(pady=5)
    ttk.Button(root, text="Generovať rozpis", command=start).pack(pady=5)

    root.mainloop()


def uloz_json(data):
    with open(JSON_SUBOR, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=4)



from openpyxl.styles import PatternFill, Font

def vytvor_excel(data, rok, mesiac, pocet_dni):
    wb = Workbook()

    # =====================
    # FARBY
    # =====================
    fill_denna = PatternFill(start_color="FFF9E79F", end_color="FFF9E79F", fill_type="solid")
    fill_nocna = PatternFill(start_color="FF85C1E9", end_color="FF85C1E9", fill_type="solid")
    bold_font = Font(bold=True)

    # =====================
    # SHEET 1 – OSOBY × DNI
    # =====================
    ws1 = wb.active
    ws1.title = "Rozpis (osoby)"

    ws1.cell(row=1, column=1, value="Meno").font = bold_font

    for den in range(1, pocet_dni + 1):
        cell = ws1.cell(row=1, column=den + 1, value=den)
        cell.font = bold_font

    for r, meno in enumerate(ZAMESTNANCI, start=2):
        ws1.cell(row=r, column=1, value=meno)

        for den in range(1, pocet_dni + 1):
            datum = f"{rok}-{mesiac:02d}-{den:02d}"

            cell = ws1.cell(row=r, column=den + 1)

            if meno in data[datum]["denna"]:
                cell.value = "D"
                cell.fill = fill_denna
            elif meno in data[datum]["nocna"]:
                cell.value = "N"
                cell.fill = fill_nocna
            else:
                cell.value = "-"

    # =====================
    # SHEET 2 – DNI × SLUŽBY
    # =====================
    ws2 = wb.create_sheet(title="Rozpis (dni)")

    ws2.cell(row=1, column=1, value="Dátum").font = bold_font
    ws2.cell(row=1, column=2, value="Typ služby").font = bold_font
    ws2.cell(row=1, column=3, value="Zamestnanci").font = bold_font

    row = 2

    for datum, sluzby in data.items():
        # Denná
        ws2.cell(row=row, column=1, value=datum)
        ws2.cell(row=row, column=2, value="Denná").fill = fill_denna
        ws2.cell(row=row, column=3, value=", ".join(sluzby["denna"]))
        row += 1

        # Nočná
        ws2.cell(row=row, column=1, value=datum)
        ws2.cell(row=row, column=2, value="Nočná").fill = fill_nocna
        ws2.cell(row=row, column=3, value=", ".join(sluzby["nocna"]))
        row += 1

    wb.save(EXCEL_SUBOR)



def vypis_terminal(data):
    print("📅 Mesačný rozpis služieb:\n")
    for datum, sluzby in data.items():
        print(f"{datum}")
        print("  Denná:", ", ".join(sluzby["denna"]))
        print("  Nočná:", ", ".join(sluzby["nocna"]))
        print()

# =====================
# MAIN
# =====================



def main():
    rok, mesiac, pocet_dni = zisti_mesiac()

    spusti_ui(pocet_dni)

    for pokus in range(1, MAX_POKUSOV + 1):
        try:
            print(f"🔄 Pokus {pokus} o generovanie rozpisu...")
            rozpis = vytvor_rozpis(rok, mesiac, pocet_dni)
            print("✅ Rozpis úspešne vygenerovaný!")
            break
        except ValueError as e:
            print(f"❌ Neúspešný pokus {pokus}: {e}")
    else:
        print("🚨 Nepodarilo sa nájsť riešenie ani po viacerých pokusoch.")
        return

    uloz_json(rozpis)
    vytvor_excel(rozpis, rok, mesiac, pocet_dni)
    vypis_terminal(rozpis)

    print("📄 JSON súbor:", JSON_SUBOR)
    print("📊 Excel súbor:", EXCEL_SUBOR)


if __name__ == "__main__":
    main()
